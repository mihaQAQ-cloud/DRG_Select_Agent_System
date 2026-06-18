"""
测试用例生成智能体（单文件实现）
生成三类用例：正常场景（程序化）、边界场景（程序化）、异常场景（LLM 或内置模板）
导出 Excel，提交到文档系统。

修复说明（v1.3）：
  1. 修复 Excel 导出错误：确保所有写入单元格的值均为字符串或基本类型，避免 openpyxl 无法处理空列表。
  2. 增强 generate_abnormal 的鲁棒性。
"""
import json
import os
import re
import shutil
import tempfile
import uuid
from typing import Optional, List, Dict, Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import requests

from agents.llm_service import LLMService

llm = LLMService()
DOC_SYSTEM_URL = "http://localhost:8001"

# ── 用例构造辅助 ──────────────────────────────────────────────────────────────

_SAMPLE_NAMES = {
    "G20": "帕金森病",
    "G35": "多发性硬化",
    "G43.9": "偏头痛",
    "G41.9": "癫痫持续状态",
    "I63.9": "脑梗死",
    "H10.3": "急性结膜炎",
    "H26.0": "老年性白内障",
    "H40.1": "开角型青光眼",
    "H52.1": "近视",
    "I10": "高血压",
    "I21.9": "急性心肌梗死",
    "I48.0": "心房颤动",
    "I50.0": "充血性心力衰竭",
    "J18.9": "肺炎",
    "J44.1": "COPD急性加重",
    "J45.9": "哮喘",
    "J96.0": "急性呼吸衰竭",
    "K25.0": "胃溃疡急性出血",
    "K57.3": "无出血的憩室病",
    "K72.0": "急性肝功能衰竭",
    "K80.0": "胆囊结石伴急性胆囊炎",
    "E11.9": "2型糖尿病",
    "E78.5": "高脂血症",
    "N18.3": "慢性肾脏病3期",
    "N17.9": "急性肾衰竭",
}


def _make_emr(
    patient_id: str,
    main_icd: str,
    secondary_icds: list,
    op_code: Optional[str],
    secondary_ops: Optional[List[str]] = None, # 新增参数
    age: int = 60,
    gender: str = "M",
) -> dict:
    return {
        "patient_id": patient_id,
        "main_diagnosis": {"icd_code": main_icd, "name": _SAMPLE_NAMES.get(main_icd, "")},
        "secondary_diagnoses": [
            {"icd_code": c, "name": _SAMPLE_NAMES.get(c, "")} for c in secondary_icds
        ],
        "main_op_code": op_code,
        "secondary_ops": secondary_ops or [], # 新增
        "age": age,
        "gender": gender,
    }


# ── CC/MCC 评估辅助（与 DRGRuleEngine.eval_cc_mcc 完全对齐）────────────────

_EXT_RE = re.compile(r'[xX]\d+$')


def _normalize_code(code: str) -> str:
    """去掉中国扩展后缀 xNNN/XNNN，保留基础 ICD 编码。"""
    return _EXT_RE.sub('', code).rstrip('.')


def _build_forward_index(raw_list: list) -> dict:
    """按 3 字符大类索引，加速前向前缀查找。"""
    idx: dict = {}
    for c in raw_list:
        base = _normalize_code(c)
        cat = base[:3]
        if cat not in idx:
            idx[cat] = []
        idx[cat].append(base)
    return idx


def _matches_forward(code: str, forward_idx: dict) -> bool:
    """规则表中是否存在以 code 为前缀的条目。"""
    norm = _normalize_code(code)
    if len(norm) < 3:
        return False
    cat = norm[:3]
    return any(entry.startswith(norm) for entry in forward_idx.get(cat, []))


def _eval_cc_mcc(
    rules: dict,
    secondary_codes: list,
    mdc_code: str,
    main_diag_code: str,
) -> str:
    """与 DRGRuleEngine.eval_cc_mcc 逻辑完全一致。"""
    excl_table: dict = rules.get("exclusion_table", {})
    excluded: set = set(excl_table.get(f"{mdc_code},{main_diag_code}", []))

    mcc_idx = _build_forward_index(rules.get("mcc_list", []))
    cc_idx = _build_forward_index(rules.get("cc_list", []))

    has_mcc = False
    has_cc = False
    for code in secondary_codes:
        if code in excluded:
            continue
        if _matches_forward(code, mcc_idx):
            has_mcc = True
            break
        if _matches_forward(code, cc_idx):
            has_cc = True

    if has_mcc:
        return "MCC"
    if has_cc:
        return "CC"
    return "无"


def _derive_expected_drg(
    rules: dict,
    main_icd: str,
    op_code: Optional[str],
    secondary_icds: list,
    secondary_ops: List[str] = [] # 新增
) -> str:
    """
    从规则推导期望 DRG 组号，与 DRGRuleEngine.classify 完全对齐。
    """
    # 1. MDC 匹配
    mdc_map = rules.get("mdc_mapping", {})
    mdc_code = None
    for prefix in sorted(mdc_map.keys(), key=len, reverse=True):
        if main_icd.upper().startswith(prefix.upper()):
            mdc_code = mdc_map[prefix]
            break
    if not mdc_code:
        return "UNKNOWN"

    # 2. ADRG 匹配
    adrg_map = rules.get("adrg_mapping", {})
    best_adrg: Optional[str] = None
    best_len: int = 0
    fallback: Optional[str] = None

    # 【修复点】收集所有有效手术编码，包括主手术和其他手术
    all_ops = []
    if op_code:
        all_ops.append(op_code)
    if secondary_ops:
        all_ops.extend([op for op in secondary_ops if op])

    for key, adrg in adrg_map.items():
        if "," not in key:
            continue
        k_mdc, k_op = key.split(",", 1)
        if k_mdc != mdc_code:
            continue
        if k_op == "none":
            fallback = adrg
            continue
        for current_op in all_ops:
            if current_op.startswith(k_op) and len(k_op) > best_len:
                best_len = len(k_op)
                best_adrg = adrg

    if best_adrg:
        adrg_code = best_adrg
    elif not all_ops:
        diag_map: dict = rules.get("diag_adrg_mapping", {})
        diag_match: Optional[str] = None
        for prefix in sorted(diag_map.keys(), key=len, reverse=True):
            if main_icd.upper().startswith(prefix.upper()):
                diag_match = diag_map[prefix]
                break
        adrg_code = diag_match if diag_match else (fallback or "??")
    else:
        adrg_code = fallback or "??"

    # 3. CC/MCC 评估
    cc_mcc = _eval_cc_mcc(rules, secondary_icds, mdc_code, main_icd)
    sev = {"MCC": "1", "CC": "3", "无": "5"}[cc_mcc]

    return adrg_code + sev


# ── 三类用例生成 ──────────────────────────────────────────────────────────────

def generate_normal(rules: dict) -> list:
    """
    遍历规则中所有有效的 MDC × ADRG 组合，程序化生成标准入组测试用例。
    """
    cases = []
    adrg_map = rules.get("adrg_mapping", {})
    sample_icd = rules.get("sample_icd_codes", {})
    sample_op = rules.get("sample_op_codes", {})

    seen_adrgs: set = set()
    for key, adrg in adrg_map.items():
        if "," not in key:
            continue
        mdc_code, op_prefix = key.split(",", 1)
        if adrg in seen_adrgs:
            continue
        seen_adrgs.add(adrg)

        icds = sample_icd.get(mdc_code, [])
        ops = sample_op.get(mdc_code, [])
        if not icds:
            continue

        main_icd = icds[0]
        op_code = None
        if op_prefix != "none" and ops:
            op_code = next((o for o in ops if o.startswith(op_prefix)), ops[0])

        case_id = f"TC-N-{len(cases) + 1:03d}"
        emr = _make_emr(case_id, main_icd, [], op_code)
        expected = _derive_expected_drg(rules, main_icd, op_code, [])

        cases.append({
            "case_id": case_id,
            "type": "normal",
            "input": emr,
            "expected_drg": expected,
            "description": f"{mdc_code} {adrg} 标准入组（无并发症）",
        })
    return cases


def generate_boundary(rules: dict) -> list:
    """
    针对 CC 有/无、MCC 有/无、排除表边界生成测试用例。
    """
    cases = []
    sample_icd = rules.get("sample_icd_codes", {})
    mcc_list = rules.get("mcc_list", [])
    cc_list = rules.get("cc_list", [])
    excl_table = rules.get("exclusion_table", {})
    sample_op = rules.get("sample_op_codes", {})

    for mdc_code, icds in sample_icd.items():
        if not icds:
            continue
        main_icd = icds[0]
        ops = sample_op.get(mdc_code, [])
        op_code = ops[0] if ops else None

        # 边界1: 带CC次诊断
        if cc_list:
            cc_code = cc_list[0]
            emr = _make_emr(f"TC-B-{len(cases)+1:03d}", main_icd, [cc_code], None)
            expected = _derive_expected_drg(rules, main_icd, None, [cc_code])
            cases.append({
                "case_id": f"TC-B-{len(cases)+1:03d}",
                "type": "boundary",
                "input": emr,
                "expected_drg": expected,
                "description": f"{mdc_code} 带CC次诊断({cc_code})边界",
            })

        # 边界2: 带MCC次诊断
        if mcc_list:
            mcc_code = mcc_list[0]
            emr = _make_emr(f"TC-B-{len(cases)+1:03d}", main_icd, [mcc_code], None)
            expected = _derive_expected_drg(rules, main_icd, None, [mcc_code])
            cases.append({
                "case_id": f"TC-B-{len(cases)+1:03d}",
                "type": "boundary",
                "input": emr,
                "expected_drg": expected,
                "description": f"{mdc_code} 带MCC次诊断({mcc_code})边界",
            })

        # 边界3: 排除表命中
        excl_key = f"{mdc_code},{main_icd}"
        if excl_key in excl_table and cc_list:
            excluded_code = excl_table[excl_key][0]
            emr = _make_emr(f"TC-B-{len(cases)+1:03d}", main_icd, [excluded_code], None)
            expected = _derive_expected_drg(rules, main_icd, None, [excluded_code])
            cases.append({
                "case_id": f"TC-B-{len(cases)+1:03d}",
                "type": "boundary",
                "input": emr,
                "expected_drg": expected,
                "description": f"{mdc_code} 排除表命中：次诊断{excluded_code}应被排除",
            })

        # 边界4: 有手术编码 vs 无手术编码
        if op_code:
            emr_surg = _make_emr(f"TC-B-{len(cases)+1:03d}", main_icd, [], op_code)
            emr_nosurg = _make_emr(f"TC-B-{len(cases)+2:03d}", main_icd, [], None)
            exp_surg = _derive_expected_drg(rules, main_icd, op_code, [])
            exp_nosurg = _derive_expected_drg(rules, main_icd, None, [])
            cases.append({
                "case_id": f"TC-B-{len(cases)+1:03d}",
                "type": "boundary",
                "input": emr_surg,
                "expected_drg": exp_surg,
                "description": f"{mdc_code} 有手术编码({op_code})→外科ADRG",
            })
            cases.append({
                "case_id": f"TC-B-{len(cases)+2:03d}",
                "type": "boundary",
                "input": emr_nosurg,
                "expected_drg": exp_nosurg,
                "description": f"{mdc_code} 无手术编码→内科ADRG",
            })

        # 边界5: 其他手术命中 ADRG
        if ops and len(ops) > 1:
            # 构造一个主手术为空，但其他手术包含关键编码的用例
            other_op = ops[1] 
            emr_other = _make_emr(f"TC-B-{len(cases)+1:03d}", main_icd, [], None, secondary_ops=[other_op])
            exp_other = _derive_expected_drg(rules, main_icd, None, [], secondary_ops=[other_op]) # 需更新 derive 函数签名
            cases.append({
                "case_id": f"TC-B-{len(cases)+1:03d}",
                "type": "boundary",
                "input": emr_other,
                "expected_drg": exp_other,
                "description": f"{mdc_code} 主手术无，但其他手术({other_op})命中 ADRG",
            })

    return cases


# 内置异常用例模板
_BUILTIN_ABNORMAL = [
    {
        "type": "abnormal",
        "input": {"patient_id": "TC-A-001", "main_diagnosis": {"icd_code": "ZZ9"}, "secondary_diagnoses": [], "main_op_code": None, "age": 50, "gender": "M", "los": 3},
        "expected_drg": "ERROR_422",
        "description": "无效ICD编码（ZZ9），期望返回422参数错误",
    },
    {
        "type": "abnormal",
        "input": {"patient_id": "TC-A-002", "main_diagnosis": {"icd_code": "999X"}, "secondary_diagnoses": [], "main_op_code": None, "age": 50, "gender": "M", "los": 3},
        "expected_drg": "ERROR_422",
        "description": "编码格式错误（999X，非字母开头），期望返回422",
    },
    {
        "type": "abnormal",
        "input": {"patient_id": "TC-A-003", "main_diagnosis": {"icd_code": "Z00.0"}, "secondary_diagnoses": [], "main_op_code": None, "age": 50, "gender": "M", "los": 3},
        "expected_drg": "ERROR_422",
        "description": "Z编码不在任何MDC范围内，期望返回422",
    },
    {
        "type": "abnormal",
        "input": {"patient_id": "TC-A-004", "main_diagnosis": {"icd_code": "J44.1"}, "secondary_diagnoses": [], "main_op_code": None, "age": -1, "gender": "M", "los": 7},
        "expected_drg": "ERROR_422",
        "description": "年龄为负数，期望返回422参数校验错误",
    },
    {
        "type": "abnormal",
        "input": {"patient_id": "TC-A-005", "main_diagnosis": {"icd_code": "J44.1"}, "secondary_diagnoses": [], "main_op_code": None, "age": 60, "gender": "X", "los": 7},
        "expected_drg": "ERROR_422",
        "description": "性别字段非M/F，期望返回422",
    },
    {
        "type": "abnormal",
        "input": {"patient_id": "TC-A-006", "main_diagnosis": {"icd_code": "J44.1"}, "secondary_diagnoses": [{"icd_code": "INVALID"}], "main_op_code": None, "age": 60, "gender": "M", "los": 7},
        "expected_drg": "ERROR_422",
        "description": "次诊断包含非法ICD编码，期望返回422",
    },
    {
        "type": "abnormal",
        "input": {"patient_id": "TC-A-007", "secondary_diagnoses": [], "main_op_code": None, "age": 60, "gender": "M", "los": 7},
        "expected_drg": "ERROR_422",
        "description": "缺少主诊断字段，期望返回422",
    },
    {
        "type": "abnormal",
        "input": {"patient_id": "TC-A-008", "main_diagnosis": {"icd_code": "J44.1"}, "secondary_diagnoses": [], "main_op_code": None, "age": 60, "gender": "M", "los": 0},
        "expected_drg": "ES35",
        "description": "住院天数为0（当天入出院），期望正常入组 ES35",
    },
    {
        "type": "abnormal",
        "input": {"patient_id": "TC-A-009", "main_diagnosis": {"icd_code": "J44.1"}, "secondary_diagnoses": [], "main_op_code": None, "age": 120, "gender": "F", "los": 365},
        "expected_drg": "ES35",
        "description": "极端年龄（120岁）和住院天数（365天），期望正常入组 ES35",
    },
    {
        "type": "abnormal",
        "input": {"patient_id": "TC-A-010", "main_diagnosis": {"icd_code": "J44.1"}, "secondary_diagnoses": [{"icd_code": f"J{i:02d}.0"} for i in range(10, 50)], "main_op_code": None, "age": 60, "gender": "M", "los": 10},
        "expected_drg": "ES31",
        "description": "次诊断超过40条（压力测试），J11.0命中MCC→期望入组ES31",
    },
]


def generate_abnormal(n: int = 10) -> List[Dict[str, Any]]:
    """
    调用 LLM 生成异常场景用例；LLM 不可用时使用内置模板。
    修复：严格校验返回值类型，确保返回字典列表。
    """
    if not llm.is_configured():
        cases = _BUILTIN_ABNORMAL[:n]
        for i, c in enumerate(cases):
            c["case_id"] = f"TC-A-{i+1:03d}"
        return cases

    prompt = (
        f"请生成 {n} 个 DRG 入组 REST 接口（POST /api/drg/classify）的异常测试用例，"
        "以 JSON 数组格式返回，每个元素包含字段：\n"
        "case_id（字符串）、type（固定为 'abnormal'）、"
        "input（请求体 JSON，包含 patient_id/main_diagnosis/secondary_diagnoses/"
        "main_op_code/age/gender/los）、"
        "expected_drg（期望结果，无法入组时填 'ERROR_422'）、"
        "description（说明）。\n"
        "涵盖：ICD 编码格式错误、字段缺失、编码不在 MDC 范围内、极端值等场景。"
        "仅返回 JSON 数组，不要添加任何解释文字。"
    )
    raw = llm.chat(prompt)
    try:
        start = raw.find("[")
        end = raw.rfind("]") + 1
        if start == -1 or end == 0:
            raise ValueError("No JSON array found in LLM response")
        
        parsed = json.loads(raw[start:end])
        
        # 严格校验：必须是列表，且元素必须是字典
        if not isinstance(parsed, list):
            raise ValueError("LLM response is not a list")
            
        valid_cases = []
        for item in parsed:
            if isinstance(item, dict):
                # 补充缺失的 case_id
                if "case_id" not in item:
                    item["case_id"] = f"TC-A-{len(valid_cases)+1:03d}"
                valid_cases.append(item)
        
        if not valid_cases:
            raise ValueError("No valid dictionary cases found in LLM response")
            
        return valid_cases[:n]

    except Exception as e:
        print(f"[TestAgent] LLM 生成异常用例失败 ({e})，降级使用内置模板")
        cases = _BUILTIN_ABNORMAL[:n]
        for i, c in enumerate(cases):
            c["case_id"] = f"TC-A-{i+1:03d}"
        return cases


# ── 运行所有用例、导出 Excel、提交文档 ────────────────────────────────────────

def _compute_coverage(rules: dict, all_cases: list) -> dict:
    """FA-03-05：统计 DRG 规则条目覆盖率。"""
    adrg_map = rules.get("adrg_mapping", {})
    adrg_set = set(adrg_map.values())
    covered: set = set()
    for case in all_cases:
        # 增加类型保护
        if not isinstance(case, dict):
            continue
        drg = case.get("expected_drg", "")
        if len(drg) >= 3 and drg not in ("UNKNOWN", "ERROR_422"):
            covered.add(drg[:-1]) 
    total = len(adrg_set)
    hit = len(covered & adrg_set)
    missing = sorted(adrg_set - covered)
    return {
        "total_adrg": total,
        "covered_adrg": hit,
        "coverage_rate": f"{hit/total*100:.1f}%" if total else "N/A",
        "missing_adrg": missing,
    }


def run_all(rules_path: str = "rules/drg_rules.json") -> list:
    """
    串联三类用例生成，导出 Excel，提交到文档系统，返回全部用例列表。
    """
    with open(rules_path, encoding="utf-8") as f:
        rules = json.load(f)

    normal = generate_normal(rules)
    boundary = generate_boundary(rules)
    abnormal = generate_abnormal(10)
    all_cases = normal + boundary + abnormal

    # 覆盖度统计
    coverage = _compute_coverage(rules, all_cases)
    print(
        f"[TestAgent] 用例总数: {len(all_cases)}（正常 {len(normal)} / "
        f"边界 {len(boundary)} / 异常 {len(abnormal)}）"
    )
    print(
        f"[TestAgent] ADRG 覆盖率: {coverage['coverage_rate']} "
        f"（{coverage['covered_adrg']}/{coverage['total_adrg']}）"
    )
    if coverage["missing_adrg"]:
        print(f"[TestAgent] 未覆盖 ADRG: {coverage['missing_adrg']}")

    # 导出 Excel
    os.makedirs("data/outputs", exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # ── 表头样式 ──────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill_main = PatternFill("solid", fgColor="2F5496")   
    header_fill_sub  = PatternFill("solid", fgColor="4472C4")   
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    INPUT_COLS = 8          
    main_headers = {
        "A1": "用例ID",
        "B1": "类型",
        "C1": "输入病历",          
        "K1": "期望DRG组号",
        "L1": "说明",
    }
    for cell_addr, val in main_headers.items():
        cell = ws[cell_addr]
        cell.value = val
        cell.font = header_font
        cell.fill = header_fill_main
        cell.alignment = center

    sub_headers = ["患者ID", "主诊断ICD", "诊断名称", "次诊断ICD列表", "手术编码", "年龄", "性别", "其他手术编码"]
    for i, h in enumerate(sub_headers):
        cell = ws.cell(row=2, column=3 + i, value=h)   
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill_sub
        cell.alignment = center

    for col in [1, 2, 11, 12]:   
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        ws.cell(row=1, column=col).alignment = center

    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=10)

    col_widths = {
        1: 12, 2: 10, 3: 14, 4: 16, 5: 16, 6: 28, 7: 14, 8: 8, 9: 8, 10: 14, 11: 28, 12: 36,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    # ── 数据行 ────────────────────────────────────────────────────────────────
    for row_idx, c in enumerate(all_cases, start=3):
        # 增加类型保护，跳过非字典元素
        if not isinstance(c, dict):
            continue
            
        inp = c.get("input", {})
        if not isinstance(inp, dict):
            inp = {}
            
        main_diag = inp.get("main_diagnosis", {})
        if not isinstance(main_diag, dict):
            main_diag = {}
            
        secondary = inp.get("secondary_diagnoses", [])
        if not isinstance(secondary, list):
            secondary = []
            
        # 【修复点】确保 sec_str 始终是字符串，即使 secondary 为空
        sec_list = []
        for d in secondary:
            if isinstance(d, dict):
                code = d.get("icd_code", "")
                name = d.get("name", "")
                if name:
                    sec_list.append(f"{code}({name})")
                else:
                    sec_list.append(code)
        
        sec_str = "；".join(sec_list) if sec_list else "无"

        # 【修改4】获取其他手术编码
        secondary_ops = inp.get("secondary_ops", [])
        if not isinstance(secondary_ops, list):
            secondary_ops = []
        sec_ops_str = "；".join(secondary_ops) if secondary_ops else "无"

        row = [
            c.get("case_id", ""),           
            c.get("type", ""),              
            inp.get("patient_id", ""),      
            main_diag.get("icd_code", ""),  
            main_diag.get("name", ""),      
            sec_str,                        
            inp.get("main_op_code") or "无",
            inp.get("age", ""),             
            inp.get("gender", ""),          
            sec_ops_str,             
            c.get("expected_drg", ""),      
            c.get("description", ""),       
        ]
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = left
            cell.font = Font(size=10)

    ws.freeze_panes = "A3"

    # 覆盖度汇总 sheet
    ws2 = wb.create_sheet("覆盖度统计")
    ws2.append(["指标", "值"])
    ws2.append(["ADRG 总数", coverage["total_adrg"]])
    ws2.append(["已覆盖 ADRG 数", coverage["covered_adrg"]])
    ws2.append(["覆盖率", coverage["coverage_rate"]])
    ws2.append(["未覆盖 ADRG", ", ".join(coverage["missing_adrg"])])

    # 写临时文件，POST 给 doc_system
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)

    # 提交到文档系统
    try:
        with open(tmp_path, "rb") as f:
            resp = requests.post(
                f"{DOC_SYSTEM_URL}/api/docs/submit",
                files={"file": ("test_cases.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"doc_type": "TestCase", "title": "测试用例集", "version": "1.0"},
                timeout=5,
            )
        doc_id = resp.json().get("doc_id", "unknown")
        print(f"[TestAgent] 测试用例已提交，doc_id={doc_id}")
    except Exception as exc:
        fallback_path = "data/outputs/test_cases.xlsx"
        shutil.copy2(tmp_path, fallback_path)
        print(f"[TestAgent] 文档系统提交失败（{exc}），已保存至 {fallback_path}")
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    return all_cases